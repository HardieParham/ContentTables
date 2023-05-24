from openpyxl.styles import PatternFill, Border, Alignment, Side, Font

from .content import Content, COLORS


class ContentFolder(Content):
    def __init__(self, name: str, level: int, path: str, src: str, dest: str) -> None:
        super().__init__(name, src, dest)
        self.level = level
        self.filetype = 'Folder'
        self.path = str(path)
        self.border = self.set_borders()
        self.font = self.set_font()
        self.fill = PatternFill(patternType='solid', bgColor=COLORS['red'], fgColor=COLORS['red'])
        self.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


    def set_borders(self) -> object:
        vertical_style = Side(style='thin', color=COLORS['black'])
        border = Border(left=vertical_style, right=vertical_style, top=vertical_style)
        return border


    def set_font(self) -> object:
        name='Arial'
        size = 12
        font_style = Font(name=name, color=COLORS['white'], size=size)
        return font_style
    



class ContentSubFolder(Content):
    def __init__(self, name: str, level: int, path: str, src: str, dest: str) -> None:
        super().__init__(name, src, dest)
        self.level = level
        self.adj_level = int(level-2)
        self.filetype = self.adj_level*'S' + ' Folder'
        self.path = str(path)
        self.border = self.set_borders()
        self.font = self.set_font()
        self.fill = PatternFill(patternType='solid', bgColor=COLORS['white'], fgColor=COLORS['white'])
        self.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


    def set_borders(self) -> object:
        vertical_style = Side(style='thin', color=COLORS['black'])
        border = Border(left=vertical_style, right=vertical_style, top=vertical_style)
        return border


    def set_font(self) -> object:
        name='Arial'
        size = 12
        if self.adj_level % 3 == 0:
            font_style = Font(name=name, color=COLORS['red'], size=size)
        elif self.adj_level % 3 == 1:
            font_style = Font(name=name, color=COLORS['blue'], size=size)
        elif self.adj_level % 3 == 2:
            font_style = Font(name=name, color=COLORS['brown'], size=size)
        # font_style = Font(name=name, color=COLORS['red'], size=size)
        return font_style