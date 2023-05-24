import openpyxl

from ..models.blank import ContentBlank


class Workbook():
    def __init__(self, template_name: str, project_name: str) -> None:
        self.name = template_name
        self.project_name = project_name
        self.new_name = f'ContentTable.xlsx'
        self.wb = openpyxl.load_workbook(f"templates/{self.name}")
        self.ws = self.wb.worksheets[0]
        self.row_counter = 27 # to start on the correct row of the template


    def save_wb(self) -> None:
        self.wb.save(f'{self.project_name}/{self.new_name}')
        print('saved')


    def apply_blank(self, bottom:bool = False) -> None:
        item = ContentBlank(bottom=bottom)
        self.ws['A' + str(self.row_counter)].border = item.border
        self.ws['A' + str(self.row_counter)].fill = item.fill

        self.ws['B' + str(self.row_counter)].border = item.border
        self.ws['B' + str(self.row_counter)].fill = item.fill

        self.ws['C' + str(self.row_counter)].border = item.border
        self.ws['C' + str(self.row_counter)].fill = item.fill

        self.ws['D' + str(self.row_counter)].border = item.border
        self.ws['D' + str(self.row_counter)].fill = item.fill
        self.row_counter += 1



    def apply_folder(self, item: object) -> None:
        self.apply_blank()
        print(f'{self.row_counter}: {item.oldname}')

        self.ws['A' + str(self.row_counter)].value = item.filetype
        self.ws['A' + str(self.row_counter)].alignment = item.alignment
        self.ws['A' + str(self.row_counter)].border = item.border
        self.ws['A' + str(self.row_counter)].fill = item.fill
        self.ws['A' + str(self.row_counter)].font = item.font

        self.ws['B' + str(self.row_counter)].value = item.oldname
        self.ws['B' + str(self.row_counter)].hyperlink = item.path
        self.ws['B' + str(self.row_counter)].alignment = item.alignment
        self.ws['B' + str(self.row_counter)].border = item.border
        self.ws['B' + str(self.row_counter)].fill = item.fill
        self.ws['B' + str(self.row_counter)].font = item.font

        self.ws['C' + str(self.row_counter)].value = item.newname
        self.ws['C' + str(self.row_counter)].alignment = item.alignment
        self.ws['C' + str(self.row_counter)].border = item.border
        self.ws['C' + str(self.row_counter)].fill = item.fill
        self.ws['C' + str(self.row_counter)].font = item.font

        self.ws['D' + str(self.row_counter)].value = 1
        self.ws['D' + str(self.row_counter)].alignment = item.alignment
        self.ws['D' + str(self.row_counter)].border = item.border
        self.ws['D' + str(self.row_counter)].fill = item.fill
        self.ws['D' + str(self.row_counter)].font = item.font
        self.row_counter += 1


    def apply_file(self, item: object) -> None:
        print(f'{self.row_counter}: {item.oldname}')
        self.ws['A' + str(self.row_counter)].value = item.filetype
        self.ws['A' + str(self.row_counter)].alignment = item.alignment
        self.ws['A' + str(self.row_counter)].border = item.border
        self.ws['A' + str(self.row_counter)].fill = item.fill
        self.ws['A' + str(self.row_counter)].font = item.font

        self.ws['B' + str(self.row_counter)].value = item.oldname
        self.ws['B' + str(self.row_counter)].hyperlink = item.path
        self.ws['B' + str(self.row_counter)].alignment = item.alignment
        self.ws['B' + str(self.row_counter)].border = item.border
        self.ws['B' + str(self.row_counter)].fill = item.fill
        self.ws['B' + str(self.row_counter)].font = item.font

        self.ws['C' + str(self.row_counter)].value = item.newname
        self.ws['C' + str(self.row_counter)].alignment = item.alignment
        self.ws['C' + str(self.row_counter)].border = item.border
        self.ws['C' + str(self.row_counter)].fill = item.fill
        self.ws['C' + str(self.row_counter)].font = item.font

        self.ws['D' + str(self.row_counter)].value = 1
        self.ws['D' + str(self.row_counter)].alignment = item.alignment
        self.ws['D' + str(self.row_counter)].border = item.border
        self.ws['D' + str(self.row_counter)].fill = item.fill
        self.ws['D' + str(self.row_counter)].font = item.font
        self.row_counter += 1
